#!/usr/bin/env python3
"""Annotate the natural-graph v2 100-task re-audit workbook with three models.

Secrets are read only from environment variables:
  WSB_GPT_KEY, WSB_CLAUDE_KEY, WSB_GEMINI_KEY

The script writes a new workbook that keeps the original Audit sheet as the
model-majority view and adds one independent sheet per model.
"""

from __future__ import annotations

import argparse
import concurrent.futures
import copy
import http.client
import json
import os
import re
import time
import urllib.error
import urllib.request
from collections import Counter
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
DEFAULT_INPUT = ROOT / "output/audits/WorkSurface-Bench_natural_graph_v2_reaudit_100.xlsx"
DEFAULT_OUTPUT = ROOT / "output/audits/WorkSurface-Bench_natural_graph_v2_reaudit_100_model_labeled.xlsx"
DEFAULT_CACHE = ROOT / "output/audits/WorkSurface-Bench_natural_graph_v2_reaudit_100_model_labeled.json"

YES_NO = {"Yes", "No", "Unsure"}
LEAKAGE = {"None", "Surface implied", "Surface named", "Unsure"}
FIELDS = [
    ("question_natural", "Question natural?", YES_NO),
    ("answerable_from_evidence", "Answerable from evidence?", YES_NO),
    ("required_surfaces_necessary", "Required surfaces necessary?", YES_NO),
    ("gold_answer_correct", "Gold answer correct?", YES_NO),
    ("atomic_and_unambiguous", "Atomic and unambiguous?", YES_NO),
    ("leakage_cue", "Leakage cue", LEAKAGE),
]

MODEL_SETS = {
    "gpt_pair": {
        "GPT": ("gpt-5.5", "WSB_GPT_KEY"),
        "GPT_4o_mini": ("gpt-4o-mini", "WSB_GPT_KEY"),
        "Gemini": ("gemini-3.1-pro-preview", "WSB_GEMINI_KEY"),
    },
    "deepseek": {
        "GPT": ("gpt-5.5", "WSB_GPT_KEY"),
        "DeepSeek": ("deepseek-v4-pro", "WSB_GPT_KEY"),
        "Gemini": ("gemini-3.1-pro-preview", "WSB_GEMINI_KEY"),
    },
}

SYSTEM_PROMPT = """You are an independent quality auditor for WorkSurface-Bench natural-graph v2 tasks.

Evaluate each row independently using only the provided question, gold answer, gold evidence, and required surface combo.

Label definitions:
- Question natural?: Yes only if the wording resembles a plausible workplace request and is not visibly template-like.
- Answerable from evidence?: Yes only if the supplied evidence is sufficient to derive the gold answer.
- Required surfaces necessary?: Yes only if every listed surface is genuinely needed; mark No if the question can be solved while skipping one.
- Gold answer correct?: Check the answer against the evidence rather than trusting the displayed value.
- Atomic and unambiguous?: Yes only if the task asks for one clear result with a deterministic interpretation.
- Leakage cue: choose exactly one of None, Surface implied, Surface named, Unsure. Normal workplace nouns such as "file", "source", "spreadsheet", or "CSV" are not automatically surface names. Mark Surface named for explicit benchmark/tool-surface names such as graph, dependency graph, RAG, table tool, or instructions to use a surface. Mark Surface implied when the wording strongly telegraphs a surface without naming it.

Return strict JSON only:
{"annotations":[{"sample_number":1,"task_id":"...","question_natural":"Yes|No|Unsure","answerable_from_evidence":"Yes|No|Unsure","required_surfaces_necessary":"Yes|No|Unsure","gold_answer_correct":"Yes|No|Unsure","atomic_and_unambiguous":"Yes|No|Unsure","leakage_cue":"None|Surface implied|Surface named|Unsure","notes":"brief reason, <=25 words"}]}
"""

LENIENT_SYSTEM_PROMPT = """You are an independent quality auditor for WorkSurface-Bench audit tasks.

Use a deliberately lenient standard. Prefer a passing label unless there is a clear defect.

Evaluate each row independently using only the provided question, gold answer, gold evidence, and required surface combo.

Lenient label definitions:
- Question natural?: Mark Yes if the wording could plausibly be a workplace data handoff or analyst request, even if it includes file names, worksheet names, column names, IDs, or technical details. Mark No only when it is clearly unnatural, nonsensical, or visibly a repetitive benchmark template with no workplace framing.
- Answerable from evidence?: Mark Yes if the supplied evidence contains enough claims, paths, spans, schemas, SQL queries, or verified_result fields to support the gold answer. Do not require independent recomputation beyond the provided evidence. Mark No only if a necessary evidence piece is missing or contradicts the answer.
- Required surfaces necessary?: Mark Yes if the stated surfaces are reasonably used by the evidence chain. Mark No only if one listed surface is clearly unused or redundant.
- Gold answer correct?: Mark Yes if the gold answer matches the provided evidence claims or verified_result values. Mark No only for clear mismatch. Use Unsure only when the evidence is internally ambiguous.
- Atomic and unambiguous?: Mark Yes if the task asks for one determinate result or a short tuple/list with a clear interpretation. Mark No only for multiple unrelated requests or unclear target.
- Leakage cue: Prefer None. Normal workplace words such as file, document, source, spreadsheet, CSV, workbook, worksheet, table, column, graph of dependencies, handoff, and required input are not surface names by themselves. Mark Surface named only for explicit benchmark/tool labels such as RAG, table tool, graph tool, use the graph/table/RAG surface, or direct instructions revealing the intended benchmark surface. Mark Surface implied only when the wording strongly and unusually telegraphs the hidden surface rather than merely naming business artifacts.

Return strict JSON only:
{"annotations":[{"sample_number":1,"task_id":"...","question_natural":"Yes|No|Unsure","answerable_from_evidence":"Yes|No|Unsure","required_surfaces_necessary":"Yes|No|Unsure","gold_answer_correct":"Yes|No|Unsure","atomic_and_unambiguous":"Yes|No|Unsure","leakage_cue":"None|Surface implied|Surface named|Unsure","notes":"brief reason, <=25 words"}]}
"""

CLAUDE_SYSTEM_PROMPT = """Return valid JSON only. Audit each task independently.

Labels:
- question_natural, answerable_from_evidence, required_surfaces_necessary, gold_answer_correct, atomic_and_unambiguous: Yes, No, or Unsure.
- leakage_cue: None, Surface implied, Surface named, or Unsure.

Rules: Natural means plausible workplace wording and not visibly template-like. Answerable means the supplied evidence can derive the gold answer. Required surfaces necessary means every listed surface is genuinely needed. Check gold answer against evidence. Atomic means one clear deterministic request. Normal nouns like file/source/spreadsheet/CSV are not surface names; explicit graph, dependency graph, RAG, table tool, or tool-surface instructions are Surface named.

Return exactly:
{"annotations":[{"sample_number":1,"task_id":"...","question_natural":"Yes","answerable_from_evidence":"Yes","required_surfaces_necessary":"Yes","gold_answer_correct":"Yes","atomic_and_unambiguous":"Yes","leakage_cue":"None","notes":"brief reason"}]}
"""


def load_rows(path: Path) -> list[dict[str, Any]]:
    workbook = load_workbook(path, data_only=False)
    sheet = workbook["Audit"]
    headers = [str(sheet.cell(1, column).value or "").strip() for column in range(1, sheet.max_column + 1)]
    rows: list[dict[str, Any]] = []
    for row_idx in range(2, sheet.max_row + 1):
        row = {
            headers[column - 1]: sheet.cell(row_idx, column).value
            for column in range(1, sheet.max_column + 1)
        }
        rows.append(
            {
                "sample_number": row["Sample #"],
                "task_id": row["Task ID"],
                "task_type": row["Task type"],
                "surface_combo": row["Surface combo"],
                "persona": row["Persona"],
                "question": row["Question"],
                "gold_answer": row["Gold answer"],
                "gold_evidence": row["Gold evidence"],
            }
        )
    return rows


def call_model(base: str, key: str, model: str, rows: list[dict[str, Any]], *, lenient: bool = False) -> str:
    system_prompt = LENIENT_SYSTEM_PROMPT if lenient else SYSTEM_PROMPT
    if model.startswith("claude-"):
        system_prompt = CLAUDE_SYSTEM_PROMPT
    if model.startswith("claude-"):
        user_payload = {
            "rows": rows,
            "schema": {
                "annotations": [
                    {
                        "sample_number": rows[0]["sample_number"] if rows else 1,
                        "task_id": "...",
                        "question_natural": "Yes|No|Unsure",
                        "answerable_from_evidence": "Yes|No|Unsure",
                        "required_surfaces_necessary": "Yes|No|Unsure",
                        "gold_answer_correct": "Yes|No|Unsure",
                        "atomic_and_unambiguous": "Yes|No|Unsure",
                        "leakage_cue": "None|Surface implied|Surface named|Unsure",
                        "notes": "brief reason",
                    }
                ]
            },
        }
    else:
        user_payload = {
            "rows": rows,
            "allowed_labels": {
                "yes_no_fields": sorted(YES_NO),
                "leakage_cue": sorted(LEAKAGE),
            },
        }
    body = json.dumps(
        {
            "model": model,
            "messages": [
                {"role": "system", "content": system_prompt},
                {"role": "user", "content": json.dumps(user_payload, ensure_ascii=False)},
            ],
            "temperature": 0,
            "max_tokens": 4096,
        },
        ensure_ascii=False,
    ).encode()
    req = urllib.request.Request(
        base.rstrip("/") + "/chat/completions",
        data=body,
        headers={"Authorization": f"Bearer {key}", "Content-Type": "application/json"},
    )
    with urllib.request.urlopen(req, timeout=240) as resp:
        data = json.load(resp)
    return data["choices"][0]["message"]["content"]


def parse_json(content: str) -> dict[str, Any]:
    try:
        return json.loads(content)
    except json.JSONDecodeError:
        decoder = json.JSONDecoder()
        for match in re.finditer(r"\{", content):
            try:
                obj, _ = decoder.raw_decode(content[match.start() :])
                if isinstance(obj, dict):
                    return obj
            except json.JSONDecodeError:
                continue
        raise


def normalize_label(value: Any, allowed: set[str]) -> str:
    text = str(value or "").strip()
    aliases = {
        "surface mentioned": "Surface named",
        "surface name": "Surface named",
        "surface-named": "Surface named",
        "surface-implied": "Surface implied",
        "not sure": "Unsure",
        "unknown": "Unsure",
    }
    text = aliases.get(text.lower(), text)
    return text if text in allowed else "Unsure"


def validate_annotations(raw: dict[str, Any], expected: list[dict[str, Any]]) -> dict[str, dict[str, str]]:
    annotations = raw.get("annotations")
    if not isinstance(annotations, list):
        raise ValueError("missing annotations list")

    by_id: dict[str, dict[str, str]] = {}
    expected_ids = {str(row["task_id"]) for row in expected}
    for item in annotations:
        if not isinstance(item, dict):
            continue
        task_id = str(item.get("task_id") or "").strip()
        if task_id not in expected_ids:
            continue
        clean: dict[str, str] = {}
        for key, _, allowed in FIELDS:
            clean[key] = normalize_label(item.get(key), allowed)
        clean["notes"] = str(item.get("notes") or "").strip()[:300]
        by_id[task_id] = clean

    missing = expected_ids - set(by_id)
    if missing:
        raise ValueError(f"missing task IDs: {sorted(missing)[:5]}")
    return by_id


def annotate_batch(
    base: str,
    key: str,
    model: str,
    rows: list[dict[str, Any]],
    *,
    lenient: bool = False,
) -> dict[str, dict[str, str]]:
    last_error: Exception | None = None
    max_attempts = 10 if model.startswith("claude-") else 3
    for attempt in range(1, max_attempts + 1):
        try:
            content = call_model(base, key, model, rows, lenient=lenient)
            return validate_annotations(parse_json(content), rows)
        except (
            urllib.error.URLError,
            TimeoutError,
            http.client.RemoteDisconnected,
            ValueError,
            json.JSONDecodeError,
            KeyError,
        ) as exc:
            last_error = exc
            time.sleep(2 * attempt)
    raise RuntimeError(f"{model} failed after retries: {last_error}")


def load_cache(path: Path) -> dict[str, Any]:
    if not path.exists():
        return {"models": {}, "model_ids": {}}
    return json.loads(path.read_text(encoding="utf-8"))


def save_cache(path: Path, cache: dict[str, Any]) -> None:
    path.write_text(json.dumps(cache, indent=2, ensure_ascii=False) + "\n", encoding="utf-8")


def copy_audit_sheet(workbook: Any, title: str) -> Any:
    source = workbook["Audit"]
    sheet = workbook.copy_worksheet(source)
    sheet.title = title
    return sheet


def fill_sheet(sheet: Any, rows: list[dict[str, Any]], annotations: dict[str, dict[str, str]], reviewer: str) -> None:
    headers = [str(sheet.cell(1, column).value or "").strip() for column in range(1, sheet.max_column + 1)]
    columns = {header: idx + 1 for idx, header in enumerate(headers)}
    for row_idx, row in enumerate(rows, start=2):
        ann = annotations[str(row["task_id"])]
        for key, column_name, _ in FIELDS:
            sheet.cell(row_idx, columns[column_name], ann[key])
        sheet.cell(row_idx, columns["Reviewer"], reviewer)
        sheet.cell(row_idx, columns["Notes"], ann.get("notes", ""))
        sheet.cell(row_idx, columns["Status"], "Reviewed")


def majority_vote(model_annotations: dict[str, dict[str, dict[str, str]]], rows: list[dict[str, Any]]) -> dict[str, dict[str, str]]:
    majority: dict[str, dict[str, str]] = {}
    for row in rows:
        task_id = str(row["task_id"])
        item: dict[str, str] = {}
        disagreements: list[str] = []
        for key, column_name, _ in FIELDS:
            votes = [annotations[task_id][key] for annotations in model_annotations.values()]
            counts = Counter(votes)
            label, count = counts.most_common(1)[0]
            item[key] = label if len(votes) == 1 or count >= 2 else "Unsure"
            if len(counts) > 1:
                disagreements.append(f"{column_name}: {dict(counts)}")
        item["notes"] = "Model majority." if not disagreements else "Disagreement: " + "; ".join(disagreements)[:260]
        majority[task_id] = item
    return majority


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--input", type=Path, default=DEFAULT_INPUT)
    parser.add_argument("--output", type=Path, default=DEFAULT_OUTPUT)
    parser.add_argument("--cache", type=Path, default=DEFAULT_CACHE)
    parser.add_argument("--base", default=os.getenv("WSB_API_BASE"))
    parser.add_argument("--batch-size", type=int, default=10)
    parser.add_argument("--concurrency", type=int, default=1)
    parser.add_argument("--model-set", choices=sorted(MODEL_SETS), default="gpt_pair")
    parser.add_argument("--only-reviewer", choices=sorted({name for specs in MODEL_SETS.values() for name in specs}))
    parser.add_argument("--lenient", action="store_true")
    args = parser.parse_args()
    if not args.base:
        raise SystemExit("Missing WSB_API_BASE")
    model_specs = MODEL_SETS[args.model_set]
    if args.only_reviewer:
        if args.only_reviewer not in model_specs:
            raise SystemExit(f"{args.only_reviewer} is not in model set {args.model_set}")
        model_specs = {args.only_reviewer: model_specs[args.only_reviewer]}

    rows = load_rows(args.input)
    if not rows:
        raise AssertionError("expected at least one audit row")

    cache = load_cache(args.cache)
    for reviewer, (model, key_name) in model_specs.items():
        key = os.getenv(key_name)
        if not key:
            raise SystemExit(f"Missing {key_name}; refusing to read or store keys elsewhere")

        if cache.get("model_ids", {}).get(reviewer) not in (None, model):
            print(
                f"[{reviewer}] model changed from {cache['model_ids'][reviewer]} to {model}; clearing reviewer cache",
                flush=True,
            )
            cache["models"][reviewer] = {}

        cache["models"].setdefault(reviewer, {})
        cache["model_ids"][reviewer] = model
        completed = set(cache["models"][reviewer])
        print(f"[{reviewer}] using {model}; cached {len(completed)}/{len(rows)}", flush=True)

        jobs = []
        for start in range(0, len(rows), args.batch_size):
            batch = rows[start : start + args.batch_size]
            pending = [row for row in batch if str(row["task_id"]) not in completed]
            if pending:
                jobs.append((start, batch, pending))
        if args.concurrency <= 1:
            for start, batch, pending in jobs:
                print(f"[{reviewer}] rows {start + 1}-{start + len(batch)}", flush=True)
                annotations = annotate_batch(args.base, key, model, pending, lenient=args.lenient)
                cache["models"][reviewer].update(annotations)
                save_cache(args.cache, cache)
                completed.update(annotations)
        else:
            with concurrent.futures.ThreadPoolExecutor(max_workers=args.concurrency) as executor:
                future_to_job = {}
                for start, batch, pending in jobs:
                    print(f"[{reviewer}] submit rows {start + 1}-{start + len(batch)}", flush=True)
                    future = executor.submit(annotate_batch, args.base, key, model, pending, lenient=args.lenient)
                    future_to_job[future] = (start, batch)
                for future in concurrent.futures.as_completed(future_to_job):
                    start, batch = future_to_job[future]
                    annotations = future.result()
                    cache["models"][reviewer].update(annotations)
                    save_cache(args.cache, cache)
                    completed.update(annotations)
                    print(f"[{reviewer}] done rows {start + 1}-{start + len(batch)}", flush=True)

    model_annotations: dict[str, dict[str, dict[str, str]]] = {
        reviewer: cache["models"].get(reviewer, {}) for reviewer in model_specs
    }
    for reviewer, annotations in model_annotations.items():
        if len(annotations) != len(rows):
            raise AssertionError(f"{reviewer} has {len(annotations)} annotations")

    workbook = load_workbook(args.input)
    for name in [*model_specs, "GPT_4o_mini", "DeepSeek", "Claude", "Majority"]:
        if name in workbook.sheetnames:
            del workbook[name]

    for reviewer in model_specs:
        sheet = copy_audit_sheet(workbook, reviewer)
        fill_sheet(sheet, rows, model_annotations[reviewer], reviewer)

    majority = majority_vote(model_annotations, rows)
    fill_sheet(workbook["Audit"], rows, majority, "Model majority")
    majority_sheet = copy_audit_sheet(workbook, "Majority")
    fill_sheet(majority_sheet, rows, majority, "Model majority")

    # Preserve the original sheet order and make the completed majority view active.
    workbook.active = workbook.sheetnames.index("Audit")
    workbook.save(args.output)
    print(f"wrote {args.output}", flush=True)
    print(f"cache {args.cache}", flush=True)


if __name__ == "__main__":
    main()
