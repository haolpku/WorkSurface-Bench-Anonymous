"""Analyze three independent 100-task WorkSurface-Bench audit workbooks."""

from __future__ import annotations

import json
import math
from collections import Counter, defaultdict
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
AUDIT_DIR = ROOT / "output/audits"
FILES = [AUDIT_DIR / f"WorkSurface-Bench_100_Annotator_{i}.xlsx" for i in (1, 2, 3)]
OUTPUT = ROOT / "results/human_audit_analysis.json"
DIMENSIONS = [
    "Question natural?",
    "Answerable from evidence?",
    "Required surfaces necessary?",
    "Gold answer correct?",
    "Atomic and unambiguous?",
    "Leakage cue",
]


def normalize(value: object) -> str:
    text = str(value or "").strip()
    key = text.lower()
    aliases = {
        "yes": "Yes",
        "no": "No",
        "unsure": "Unsure",
        "none": "None",
        "surface named": "Surface named",
        "surface mentioned": "Surface named",
        "answer leaked": "Answer leaked",
        "other": "Other",
    }
    return aliases.get(key, text)


def load_rows(path: Path) -> list[dict]:
    workbook = load_workbook(path, read_only=True, data_only=False)
    sheet = workbook["Audit"] if "Audit" in workbook.sheetnames else workbook[workbook.sheetnames[0]]
    rows = list(sheet.iter_rows(min_row=1, max_row=101, min_col=1, max_col=21, values_only=True))
    headers = [str(value or "").strip() for value in rows[0]]
    output = []
    for values in rows[1:]:
        row = dict(zip(headers, values))
        row["Task ID"] = str(row["Task ID"]).strip()
        for dimension in DIMENSIONS:
            row[dimension] = normalize(row[dimension])
        row["Status"] = normalize(row.get("Status"))
        output.append(row)
    return output


def fleiss_kappa(label_sets: list[list[str]]) -> tuple[float, float | None]:
    n_raters = len(label_sets[0])
    totals = Counter(label for labels in label_sets for label in labels)
    observed = sum(
        (sum(count * count for count in Counter(labels).values()) - n_raters)
        / (n_raters * (n_raters - 1))
        for labels in label_sets
    ) / len(label_sets)
    expected = sum((count / (len(label_sets) * n_raters)) ** 2 for count in totals.values())
    if math.isclose(expected, 1.0):
        return observed, None
    return observed, (observed - expected) / (1 - expected)


def main() -> None:
    annotators = [load_rows(path) for path in FILES]
    mappings = [{row["Task ID"]: row for row in rows} for rows in annotators]
    ids = [row["Task ID"] for row in annotators[0]]
    if len(ids) != 100 or len(set(ids)) != 100:
        raise AssertionError("expected 100 unique task IDs")
    if any(set(mapping) != set(ids) for mapping in mappings):
        raise AssertionError("annotator task-ID sets differ")

    analysis = {
        "n_tasks": len(ids),
        "n_annotators": len(annotators),
        "source_files": [str(path) for path in FILES],
        "dimensions": {},
        "task_type_breakdown": {},
        "tasks": [],
    }
    for dimension in DIMENSIONS:
        label_sets = [[mapping[task_id][dimension] for mapping in mappings] for task_id in ids]
        agreement, kappa = fleiss_kappa(label_sets)
        majority = Counter()
        unanimous = 0
        no_majority = 0
        for labels in label_sets:
            counts = Counter(labels)
            top_label, top_count = counts.most_common(1)[0]
            if top_count >= 2:
                majority[top_label] += 1
            else:
                no_majority += 1
            unanimous += len(counts) == 1
        analysis["dimensions"][dimension] = {
            "raw_agreement": agreement,
            "fleiss_kappa": kappa,
            "unanimous_n": unanimous,
            "majority_counts": dict(majority),
            "no_majority_n": no_majority,
        }

    type_counts = defaultdict(lambda: defaultdict(Counter))
    for task_id in ids:
        base = mappings[0][task_id]
        record = {
            "task_id": task_id,
            "sample_number": base["Sample #"],
            "task_type": base["Task type"],
            "surface_combo": base["Surface combo"],
            "question": base["Question"],
            "annotators": {},
            "majority": {},
            "unanimous": {},
        }
        for index, mapping in enumerate(mappings, 1):
            record["annotators"][str(index)] = {
                dimension: mapping[task_id][dimension] for dimension in DIMENSIONS
            }
        for dimension in DIMENSIONS:
            labels = [mapping[task_id][dimension] for mapping in mappings]
            counts = Counter(labels)
            top_label, top_count = counts.most_common(1)[0]
            record["majority"][dimension] = top_label if top_count >= 2 else None
            record["unanimous"][dimension] = len(counts) == 1
            type_counts[base["Surface combo"]][dimension][record["majority"][dimension]] += 1
        analysis["tasks"].append(record)

    for combo, dimensions in sorted(type_counts.items()):
        analysis["task_type_breakdown"][combo] = {
            dimension: dict(counts) for dimension, counts in dimensions.items()
        }

    OUTPUT.write_text(json.dumps(analysis, indent=2, ensure_ascii=False) + "\n", encoding="utf-8")
    print(OUTPUT)


if __name__ == "__main__":
    main()
